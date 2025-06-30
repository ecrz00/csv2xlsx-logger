'''
Script Name:   ibuttons.py
Description:   Builds a macro from csv data
Author:        Itzel Torres (https://github.com/itzelbtl/Ibuttons_funciones)
Updated by:    Erick Gómez
Date Created: August 2023
Last Modified: June 2025
Version:       2.1.0

Overview: 

This program processes temperature data captured by iButtons and stored in CSV files, generating a consolidated .xlsx file. Each temperature measurement is timestamped with the corresponding date and time it was recorded.
The generated Excel file includes:
    -The raw data (timestamp and temperature) exactly as found in the original CSV files.
    -The processed data, where only the temperature values are grouped by experimental day.
    -All data is writen within the same xlsx file, separate by sheets.
    -An additional sheet is created to save commands(=AVERAGE(...)) to compute the average depending on the days specified into the config file.

Key features:
    -The starting point of the first experimental day is defined by the startRow parameter in the JSON configuration file. This indicates the visual row in the CSV where the first experimental day begins.
    -The endRow parameter specifies the last row of interest in the dataset.
    -If the RepeatLastValue flag is set to true, each experimental day will end with the same value that starts the following day. If false, each day's data ends normally without repetition.
    -To compute the average, it is required to specify the days to be processed. This is fone using the DaysToProcess parameter in the JSON configuration fie:
        -Set "All" to compute using all experimental days.
        -Set "None" to skip average cpmputation. 
        -Set a list of numbers, e.g., [1,2,..., experimental days], to specify which days to include. Any number can be used as long as it does not exceed the total of experimental days.  

Data validation:
    -All CSV files must have the same sample rate.
    -All CSV files must use the same temperature units (e.g., °C or °F).

The script is designed to be portable and can be placed in any folder alongside the corresponding CSV files to process them automatically.
'''
#-------------------- libraries used --------------------
import os 
import csv
import re
import json
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
from time import sleep
import sys
from datetime import datetime, timedelta

class MacroHandler():
    def __init__(self):
        self.startRow = None
        self.endRow = None
        self.repeatLastValue = None
        self.computeAverage = True
        self.validateSampleRate = []
        self.validateUnits = []
        self.sampleRate = None
        self.units = None
        self.daysToProcess = None
        self.recordsPerDay = None
        self.banner = ''' 
██╗██████╗ ██╗   ██╗████████╗████████╗ ██████╗ ███╗   ██╗███████╗
██║██╔══██╗██║   ██║╚══██╔══╝╚══██╔══╝██╔═══██╗████╗  ██║██╔════╝
██║██████╔╝██║   ██║   ██║      ██║   ██║   ██║██╔██╗ ██║███████╗
██║██╔══██╗██║   ██║   ██║      ██║   ██║   ██║██║╚██╗██║╚════██║
██║██████╔╝╚██████╔╝   ██║      ██║   ╚██████╔╝██║ ╚████║███████║
╚═╝╚═════╝  ╚═════╝    ╚═╝      ╚═╝    ╚═════╝ ╚═╝  ╚═══╝╚══════
        '''
        if getattr(sys, 'frozen', False):
            self.path = os.path.dirname(sys.executable)
        else:
            self.path = os.path.dirname(os.path.abspath(__file__))
    #end def
    def printBanner(self):
        print(self.banner)
    #-------------------- retrieves all parameters from config file --------------------
    def get_config(self):
        config_path = os.path.join(self.path, 'ibuttons.config') #use the path where the script is located and join it with the config file
        try:
            with open(config_path, 'r') as f:
                config = json.load(f) #save all parameters in config
            required_keys = ["StartRow", "EndRow", "RepeatLastValue", "DaysToProcess"] #all necessary information
            for key in required_keys:
                if key not in config:
                    raise KeyError(f"Missing required key '{key}' in config file.")
            # validates data types
            if not isinstance(config["StartRow"], int):
                raise TypeError("StartRow must be an integer.")
            if not isinstance(config["EndRow"], int):
                raise TypeError("EndRow must be an integer.")
            if not isinstance(config["RepeatLastValue"], bool):
                raise TypeError("RepeatLastValue must be a boolean.")
            
            if not isinstance(config['DaysToProcess'], str):
                if not isinstance(config["DaysToProcess"], list):
                    raise TypeError('DaysToProcess must be a list of integers or a string.')
                if not all(isinstance(day, int) and day > 0 for day in config["DaysToProcess"]):
                    raise TypeError("DaysToRepeat numbers must be positive integers")
            
            #save values into attributes
            self.startRow = config.get("StartRow", 0) - 1 #retreive startRow value, is necesarry to substrac 1 to match the csv files row
            self.endRow = config.get("EndRow", 0) - 1 #retreive endRow value, is necesarry to substrac 1 to match the csv files row
            self.repeatLastValue = config["RepeatLastValue"] #retreive RepeatLastValue value
            self.daysToProcess = config["DaysToProcess"]
        except (FileNotFoundError, json.JSONDecodeError):
            raise FileExistsError("Couldn't open or parse config file.")
        except Exception as e:
            raise RuntimeError(f"Config validation error: {e}")
    #end def
        
    #-------------------- retrieves sample rate and units from csv file --------------------
    def get_sample_rate(self, reader):
        for row in reader: #search in every full-row
            for cell in row: #search in every individual element inside the row
                if "Sample Rate" in cell:
                    match = re.search(r'(\d+)', cell)
                    if match:
                        return int(match.group(1)) #return sample rate as integer
        raise ValueError("Sample rate not found")
    #end def
    
    def get_units(self, reader):
        pattern = re.compile(r'[-+]?\d+(?:\.\d+)?\s*(°[CF])') 
        for row in reader: #search in every full-row
            for cell in row: #search in every individual element inside the row
                if "High Temperature Alarm:" in cell or "Low Temperature Alarm:" in cell:
                    match = pattern.search(cell)
                    if match:
                        return match.group(1) #return units 
        raise ValueError("Units not found")
    #end def
    
    def extract_time(self, datetime_str: str) -> str:
        try:
            return datetime_str.strip().split()[1]
        except IndexError:
            return datetime_str
    #end def
    
    def get_next_time(self,current_time_str):
        try:
            current_time = datetime.strptime(current_time_str, "%H:%M")
            next_time = current_time + timedelta(minutes=self.sampleRate)
            return next_time.strftime("%H:%M")
        except ValueError:
            raise ValueError(f"Invalid time format: '{current_time_str}'.")
    
    # -------------------- retreives all the information needed (and specified ) from csv file --------------------
    def get_data_from_csv(self, full_path: str):
        clean_data = []
        #opens csv file located in the full_path, newline assures a propper reading with EOL, latin1 is used for a special characters
        with open(full_path, newline='', encoding='latin1') as csvfile: 
            lines = list(csv.reader(csvfile, delimiter=',')) #reads all the content using coma as delimiter, convert object into a list
        sample_rate = self.get_sample_rate(lines) #retreive the sample rate per csv file
        self.validateSampleRate.append(sample_rate) #save every sample rate value
        units = self.get_units(lines) #retreive the sample rate per csv file
        self.validateUnits.append(units) #save every sample rate value
        rows_to_process = lines[self.startRow:self.endRow + 1]
        for row in rows_to_process:
            #only rows within the margin [startRow, endRow] are processed 
            try:
                timestamp = row[0] #retreive date/time value 
                value = float(row[2]) #retreive numeric value
                clean_data.append([timestamp, value]) #save information into a clean list
            except Exception as e:
                print(f'An error happened with row {row}: {e}')
        return clean_data #return the list with processed data
    
    #-------------------- from information retreive, build the excel file --------------------   
    def build_xlsx_file(self, dicti: dict, reference_time: str):
        try:
            excel_filename = os.path.join(self.path, 'processedData.xlsx') #excel file's name. Just one file with various sheets
            if os.path.exists(excel_filename): #verify files existance
                wb = load_workbook(excel_filename)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                wb.remove(ws) #remove the default sheet created beacuase other sheets will be created on the fly.
            for key, value in tqdm(dicti.items(), desc="Building and filling Excel sheets"):
                sheet_name = str(key)
                if sheet_name in wb.sheetnames: #verify if the sheet already exists
                    excel_sheet_per_subject = wb[sheet_name]
                    excel_sheet_per_subject.delete_rows(1, excel_sheet_per_subject.max_row) #clear all the content to overwrite
                else:    
                    excel_sheet_per_subject = wb.create_sheet(title=sheet_name) #create the sheet
                self.save_raw_data(excel_sheet = excel_sheet_per_subject, data_list = value) #save the raw data
                self.save_by_experimental_days(excel_sheet = excel_sheet_per_subject, data_list = value) #save the processed values
                sleep(0.1)
            if self.computeAverage:
                GA_sheet_name = 'Group Average' #sheet for the computed average
                if GA_sheet_name in wb.sheetnames:#verify if the sheet already exists
                    GA_sheet = wb[GA_sheet_name]
                    GA_sheet.delete_rows(1, GA_sheet.max_row) #clear all the content to overwrite
                else:
                    GA_sheet = wb.create_sheet(title='Group Average') #create the sheet
                self.register_time(excel_sheet=GA_sheet, reference_write=reference_time)
                for i, key in tqdm(enumerate(dicti.keys(), start=1), total=len(dicti), desc='Calculating average per subject'): #uses key to name the column 
                    self.compute_average_per_subject(excel_sheet=GA_sheet, title_per_column=key, start_column=i+1)
                    sleep(0.1)
                self.compute_average_per_hour(excel_sheet=GA_sheet) #computes average using the data obtained from compute_average_per_subject method
            wb.save(excel_filename)
        except Exception as e:
            raise Exception(e)            
    #end def 
    
    #-------------------- takes raw data and saves it --------------------
    def save_raw_data(self, excel_sheet, data_list):
        excel_sheet.append(["Original CSV values"]) #first row with this string
        excel_sheet.append([f'Units: {self.units}']) #second row with the units
        excel_sheet.append(['Date/Time', 'Value']) #third row with the title of every column
        for line in data_list:
            excel_sheet.append(line) #append every value in the list to the xsxl file
    #end def
    
    #-------------------- takes raw data, separe it into experimental days and then saves only the measurements in a specific order--------------------
    def save_by_experimental_days(self, excel_sheet, data_list):       
        records_per_day = int((24*60)/self.sampleRate) #calculate how much measurements are done per day considering the sample rate
        self.recordsPerDay = records_per_day
        chunks = []
        for i in range(0, len(data_list), records_per_day):
            chunks.append(data_list[i:i+records_per_day]) #creates a list with data devided by blocks with the respective amount of records per day 
        if self.repeatLastValue: #if repeatLastValue is true within the config file, the last value in every block is the first of the next one
            for chunk in range(1, len(chunks)):
                extra_data = chunks[chunk][0]
                chunks[chunk-1].append(extra_data)
        start_col = excel_sheet.max_column + 2 #max_column represents the first available column to use, + 2 leaves a blank column between raw and processed data
        excel_sheet.cell(row=1, column=start_col, value = 'Processed values') #header for this new section
        #iterate over every block
        for day_index, chunk in enumerate(chunks, start=1):
            col = start_col + (day_index-1)
            excel_sheet.cell(row = 3, column=col, value = f'Day {day_index}') #writes Day N as header
            #every measurement is written from row 4 onwards
            for row_offset, value in enumerate(chunk):
                excel_sheet.cell(row = 4+row_offset, column=col, value = value[1]) #value[1] is the temperature
    #end def
    
    def register_time(self, excel_sheet, reference_write: str):
        excel_sheet.cell(row=1, column = 1, value = 'Time')
        time = reference_write
        for i in range(1,self.recordsPerDay+1):
            excel_sheet.cell(row=i+1, column = 1, value = time)
            time = self.get_next_time(time)
    
    ''' every column needs this format =AVERAGE(sheet_name!D4,sheet_name!E4,sheet_name!F4,sheet_name!G4,sheet_name!H4, ...)
        Sheet's name is giving by title_per_column -> =AVERAGE(title_per_column!D4,title_per_column!E4,title_per_column!F4,title_per_column!G4,title_per_column!H4, ...)
        Numbers within self.DaysToProcess represents the column letter, day 1 is D, day 2 is E, and so on, thats why the method  convert_number_to_excel_column is used with offset = 3'''
    def compute_average_per_subject(self, excel_sheet, title_per_column ,start_column):
        excel_sheet.cell(row=1, column = start_column, value = title_per_column)
        for i in range(4, self.recordsPerDay + 4):
            formula = self.build_average_formula_per_subject(self.daysToProcess, title_per_column=title_per_column, row=i)
            excel_sheet.cell(row=i-2, column = start_column, value = formula)
    #end def
    
    ''' every row in the same column needs this format =AVERAGE(A2:D2)
        The number is within the range between the first row with a value (always 2) till the last value in the row (depends on records per day)
        First letter is always A, last letter depends on the last column with values, thats why the method convert_number_to_excel_column is used with offset = 0 using max_column as a parameter'''
    def compute_average_per_hour(self, excel_sheet):
        start_column = excel_sheet.max_column+1
        excel_sheet.cell(row = 1, column = start_column, value = 'Group average')
        last_row = excel_sheet.max_row
        for i in tqdm(range(2,last_row+1), desc='Calculating average per hour'):
            formula = self.build_average_formula_per_hour(i, last_column=start_column)
            excel_sheet.cell(row = i, column = start_column, value = formula)
            sleep(0.005)
    #end def
    
    def convert_number_to_excel_column(self, n, offset):
        n += offset
        result = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26) # Convert the number to base-26 ('cause Z -> 26), adjusting by -1 because Excel columns start at 1 (not 0)
            result = chr(65 + remainder) + result # Convert the remainder to a corresponding ASCII uppercase letter (A=65) and prepend it
        return result
    #end def
    
    #-------------------- builds the formula depending if the average computed is per sheet(subject) or per hour (using the average per subject already computed)
    def build_average_formula_per_subject(self,numbers, title_per_column, row, offset=3):
        columns = [f"{title_per_column}!{self.convert_number_to_excel_column(n, offset)}{row}" for n in numbers]
        return f"=AVERAGE({','.join(columns)})"
    #end def
    
    def build_average_formula_per_hour(self, idx, last_column):
        letter = self.convert_number_to_excel_column(last_column-1, offset=0)
        rows = f"=AVERAGE(A{idx}:{letter}{idx})"
        return rows
    #end def
    
    def main(self):
        self.printBanner() #prints banner
        self.get_config() #gets parameters from config file and validates them
        list_with_all_data = {} #dictionary to store all the data
        files_in_path = os.listdir(self.path) #list all files within the path
        csv_files = [file for file in files_in_path if file.lower().endswith('.csv')]
        for file in tqdm(csv_files, desc="Reading CSV files"):
            full_path = os.path.join(self.path, file)  # create the full path
            data = self.get_data_from_csv(full_path=full_path)  # retrieve data
            file_key = os.path.splitext(file)[0]  # remove .csv extension
            list_with_all_data[file_key] = data  # store in dict
            sleep(0.1)
        if len(set(self.validateSampleRate)) == 1: #evaluate if all the sample rates are the same
            self.sampleRate = self.validateSampleRate[0]
        else:
            raise ValueError(f'Different sample rates found in csv files')
        if len(set(self.validateUnits)) == 1: #evaluates all the units are the same
            self.units = self.validateUnits[0]
        else:
            raise ValueError(f'Different units found in csv files: {self.validateUnits}')
        total_records = self.endRow - self.startRow + 1 #estimates total records requested in config file
        records_per_day = int((24 * 60) / self.sampleRate) 
        total_days = (total_records + records_per_day - 1) // records_per_day #estimates how many experimental days are included within total_records
        if not isinstance(self.daysToProcess, str):    #verify if a string was written in the config file
            if not all(day <= total_days for day in self.daysToProcess): #verify if any number specified in DaysToProcess parameter is higher than total_days
                raise ValueError(f"One or more values in daysToReport exceed the number of experimental days ({total_days}).")
        else: #verify if the string is an allowed command
            if self.daysToProcess.lower() == 'all':
                self.daysToProcess = list(range(1, total_days+1)) #includes all the experimental days
            elif self.daysToProcess.lower() == 'none':
                self.computeAverage = False #skips average computation
            else:
                raise ValueError(f'Wrong command! daysToProcess must be "All" to compute average with all days, or "None" if no average is required.')
        start_time = self.extract_time(list_with_all_data[list(list_with_all_data.keys())[0]][0][0]) #gets date/time from the first row, then extract only the hour    
        self.build_xlsx_file(dicti=list_with_all_data, reference_time=start_time) #creates and writes in the xsxl file
    #end def
    
    
def main():
    macro = MacroHandler()
    macro.main()

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print('--- Caught Exception ---')
        print(e)
        print('----------------------------')
        input("Press any key to close the program...") #adds a time window for users to read any exception message
    finally:
        sys.exit